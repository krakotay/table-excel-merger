import asyncio
import logging
import tempfile
from typing import List, Dict, Any
from aiogram import Bot, Dispatcher
from aiogram.types import Message
from aiogram.filters import Command
import os
from dotenv import load_dotenv
from aiogram import F, types
from aiogram_media_group import media_group_handler
from inn_check import check_by_inn, merge_excel
from aiogram.types import FSInputFile
from aiogram.enums.chat_action import ChatAction

load_dotenv()


BOT_TOKEN = os.getenv("TOKEN")


# Настройка логирования
logging.basicConfig(level=logging.INFO)

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

async def process_messages(messages: List[types.Message]):
    temp_files_paths = []
    for msg in messages:
        bot_file_info = await bot.get_file(msg.document.file_id)
        file_path = bot_file_info.file_path
        f_name = msg.document.file_name
        if not os.path.exists("temp"):
            os.makedirs("temp")

        # Используем стандартный временный каталог системы вместо "temp"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir="temp") as temp_file:
            temp_file_path = temp_file.name
            logging.info(f"Создание временного файла: {temp_file_path} для '{f_name}'")
            await bot.download_file(file_path, destination=temp_file_path)
            logging.info(f"Файл '{f_name}' скачан во временный файл: {temp_file_path}")
            temp_files_paths.append({'path': temp_file_path, 'name': f_name})
    df = merge_excel(temp_files_paths[0]['path'], temp_files_paths[1]['path'] if len(temp_files_paths) > 1 else None)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", prefix="Финал_", delete=False, dir="temp") as temp_file:
        temp_file_path = temp_file.name
        df.write_excel(temp_file_path)
        await messages[0].answer_document(FSInputFile(path=temp_file_path))
    await bot.send_chat_action(messages[0].chat.id, action=ChatAction.TYPING)
    wait_time_s = df.height / 1.6
    await messages[0].answer(f"Поиск ФИО, пожалуйста, подождите... Ожидаемое время: {wait_time_s} секунд")
    new_df = check_by_inn(df) 
    COLUMNS = ["ИНН", "ФИО"]
    new_df = new_df.select(COLUMNS + [col for col in df.columns if col not in COLUMNS])
    with tempfile.NamedTemporaryFile(suffix=".xlsx", prefix="Финал_ФИО_", delete=False, dir="temp") as temp_file:
        temp_file_path = temp_file.name
        new_df.write_excel(temp_file_path)
        
        # Форматирование Excel-файла
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Открываем файл через openpyxl
            wb = openpyxl.load_workbook(temp_file_path)
            ws = wb.active
            
            # Создаем стиль границы для ячеек
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Делаем заголовки жирными, с центрированием и увеличиваем высоту в 3 раза
            header_height = 0
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                header_height = max(header_height, 12)  # базовая высота для заголовка
            
            # Устанавливаем увеличенную в 3 раза высоту для заголовков
            ws.row_dimensions[1].height = header_height * 3
            
            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:  # noqa: E722
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Автоподбор высоты строк и добавление границ для всех ячеек
            for i, row in enumerate(ws.iter_rows(), 1):
                if i > 1:  # Пропускаем первую строку (заголовки), так как уже обработали
                    max_height = 12  # минимальная высота строки
                    for cell in row:
                        # Добавляем границы для всех ячеек
                        cell.border = thin_border
                        if cell.value:
                            lines = str(cell.value).count('\n') + 1
                            max_height = max(max_height, 14 * lines)
                    ws.row_dimensions[i].height = max_height
            
            # Сохраняем отформатированный файл
            wb.save(temp_file_path)
            logging.info(f"Excel-файл {temp_file_path} успешно отформатирован")
        except Exception as e:
            logging.error(f"Ошибка при форматировании Excel-файла: {e}")
        
        await messages[0].answer_document(FSInputFile(path=temp_file_path))
    await messages[0].answer(f"Файлы '{temp_files_paths[0]['name']}' '{temp_files_paths[1]['name'] if len(temp_files_paths) > 1 else ''}' успешно скачаны и обработаны.")



@dp.message(F.document & F.document.file_name.endswith(".xlsx"), F.media_group_id)
@media_group_handler
async def excel_handler(messages: List[types.Message]):
    """Обрабатывает ровно два файла Excel, отправленные как медиагруппа."""
    if not messages[0].document:
        await messages[0].answer("Не удалось получить информацию о документе.")
        return
    if len(messages) != 2:
        await messages[0].answer("Пожалуйста, отправьте ровно два файла Excel одним сообщением (как группу).")
        return
    media_group_id = messages[0].media_group_id
    try:
        # Скачиваем оба файла
        await process_messages(messages)

    except Exception as e:
        logging.error(f"Ошибка при обработке группы {media_group_id}: {e}", exc_info=True)
        await messages[0].answer(f"Произошла ошибка при скачивании или обработке файлов: {e}")


@dp.message(F.document & F.document.file_name.endswith(".xlsx"))
async def excel_onefile_handler(message: Message):
    """Обрабатывает ровно один файл Excel."""
    if not message.document:
        await message.answer("Не удалось получить информацию о документе.")
        return
    try:
        # Скачиваем оба файла
        await process_messages([message])

    except Exception as e:
        logging.error(f"Ошибка {e}", exc_info=True)
        await message.answer(f"Произошла ошибка при скачивании или обработке файлов: {e}")


@dp.message(Command("start"))
async def start(message: Message):
    await message.answer("Добро пожаловать!")




async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
